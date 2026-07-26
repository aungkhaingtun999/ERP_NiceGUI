# ==============================================================================
# reports/pricing_report_excel.py
# ERP ENTERPRISE PRICING EXCEL REPORT ENGINE v4.0
# Product + Category + Global Markup Analysis
# Settings Controlled Pricing
# ==============================================================================


from io import BytesIO
from datetime import datetime

from openpyxl import Workbook

from openpyxl.styles import (
    Font,
    Alignment,
    Border,
    Side,
    PatternFill
)

from openpyxl.utils import get_column_letter



# ==============================================================================
# STYLE
# ==============================================================================


thin_border = Border(

    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")

)


header_font = Font(
    bold=True
)


title_font = Font(
    bold=True,
    size=16
)


center = Alignment(
    horizontal="center",
    vertical="center"
)


header_fill = PatternFill(
    fill_type="solid",
    fgColor="DDDDDD"
)



# ==============================================================================
# SAFE VALUE
# ==============================================================================

def safe_value(
    data,
    key,
    default=0
):

    value = data.get(
        key,
        default
    )

    if value is None:
        return default

    return value




# ==============================================================================
# CREATE EXCEL REPORT
# ==============================================================================


def create_pricing_excel_report(

    products,

    company_name="MYANMAR ERP"

):


    wb = Workbook()


    ws = wb.active


    ws.title = "Pricing Report"



    # ==========================================================================
    # TITLE
    # ==========================================================================


    ws.merge_cells(
        "A1:O1"
    )


    ws["A1"] = (

        company_name

        +

        " - PRODUCT PRICING REPORT"

    )


    ws["A1"].font = title_font


    ws["A1"].alignment = center



    ws.merge_cells(
        "A2:O2"
    )


    ws["A2"] = (

        "Generated : "

        +

        datetime.now()
        .strftime(
            "%Y-%m-%d %H:%M:%S"
        )

    )




    # ==========================================================================
    # HEADER
    # ==========================================================================


    headers = [

        "No",

        "Product",

        "SKU",

        "Category",

        "Cost",

        "Product Markup %",

        "Category Markup %",

        "Global Markup %",

        "Applied Source",

        "Final Markup %",

        "Expected Selling",

        "Actual Selling",

        "Difference",

        "Profit",

        "Status"

    ]



    row = 4



    for col, header in enumerate(

        headers,

        start=1

    ):


        cell = ws.cell(

            row=row,

            column=col

        )


        cell.value = header


        cell.font = header_font


        cell.alignment = center


        cell.border = thin_border


        cell.fill = header_fill





    # ==========================================================================
    # DATA
    # ==========================================================================


    row += 1


    total_profit = 0


    total_sales = 0



    for index, product in enumerate(

        products,

        start=1

    ):



        cost = float(

            safe_value(

                product,

                "cost",

                product.get(

                    "purchase_price",

                    0

                )

            )

        )



        selling = float(

            safe_value(

                product,

                "actual_selling_price",

                product.get(

                    "selling_price",

                    0

                )

            )

        )



        expected = float(

            safe_value(

                product,

                "expected_selling_price",

                0

            )

        )



        profit = float(

            safe_value(

                product,

                "profit",

                selling - cost

            )

        )



        difference = float(

            safe_value(

                product,

                "price_difference",

                selling - expected

            )

        )



        total_profit += profit


        total_sales += selling



        status = (

            "OK"

            if difference >= 0

            else

            "Below Expected"

        )





        data = [

            index,


            product.get(
                "name",
                ""
            ),


            product.get(
                "sku",
                ""
            ),


            product.get(
                "category",
                "-"
            ),


            cost,


            product.get(
                "product_markup"
            ),


            product.get(
                "category_markup"
            ),


            product.get(
                "global_markup"
            ),


            product.get(
                "markup_source",
                "GLOBAL_DEFAULT_MARKUP"
            ),


            product.get(
                "final_markup_percent",
                0
            ),


            expected,


            selling,


            difference,


            profit,


            status

        ]



        for col, value in enumerate(

            data,

            start=1

        ):


            cell = ws.cell(

                row=row,

                column=col

            )


            cell.value = value


            cell.border = thin_border



            if col in [

                5,

                11,

                12,

                13,

                14

            ]:


                cell.number_format = (

                    '#,##0.00'

                )



            if col in [

                6,

                7,

                8,

                10

            ]:


                cell.number_format = (

                    '0.00"%"'

                )



        row += 1





    # ==========================================================================
    # SUMMARY
    # ==========================================================================


    row += 2



    summary = [

        (

            "Total Products",

            len(products)

        ),


        (

            "Total Selling Value",

            total_sales

        ),


        (

            "Total Profit",

            total_profit

        )

    ]



    for label, value in summary:


        ws.cell(

            row=row,

            column=1

        ).value = label



        ws.cell(

            row=row,

            column=2

        ).value = value



        if isinstance(

            value,

            float

        ):


            ws.cell(

                row=row,

                column=2

            ).number_format = (

                '#,##0.00'

            )


        row += 1





    # ==========================================================================
    # AUTO WIDTH
    # ==========================================================================


    for column in ws.columns:


        max_length = 0


        column_letter = get_column_letter(

            column[0].column

        )


        for cell in column:


            try:

                length = len(

                    str(

                        cell.value

                    )

                )


                if length > max_length:

                    max_length = length


            except Exception:

                pass



        ws.column_dimensions[

            column_letter

        ].width = max_length + 4





    # ==========================================================================
    # RETURN FILE
    # ==========================================================================


    output = BytesIO()


    wb.save(

        output

    )


    output.seek(

        0

    )


    return output
