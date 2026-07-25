# ==============================================================================
# reports/pricing_report_excel.py
# ERP ENTERPRISE PRICING EXCEL REPORT ENGINE v1.0
# ==============================================================================


from io import BytesIO

from datetime import datetime


from openpyxl import Workbook


from openpyxl.styles import (
    Font,
    Alignment,
    Border,
    Side
)


from openpyxl.utils import (
    get_column_letter
)



# ==============================================================================
# STYLE
# ==============================================================================


thin_border = Border(

    left=Side(style="thin"),

    right=Side(style="thin"),

    top=Side(style="thin"),

    bottom=Side(style="thin")

)



# ==============================================================================
# CREATE EXCEL REPORT
# ==============================================================================


def create_pricing_excel_report(
    products,
    company_name="MYANMAR ERP"
):

    """
    Generate Product Pricing Report Excel

    products:

    [
        {
            "name":"Earphones",
            "sku":"EP001",
            "category":"Personal Care",
            "purchase_price":7000,
            "markup_percent":15,
            "selling_price":8050
        }
    ]

    """


    wb = Workbook()


    ws = wb.active


    ws.title = "Pricing Report"



    # ==========================================================================
    # TITLE
    # ==========================================================================

    ws.merge_cells("A1:H1")


    title = ws["A1"]

    title.value = (
        company_name
        +
        " - PRODUCT PRICING REPORT"
    )


    title.font = Font(
        bold=True,
        size=16
    )


    title.alignment = Alignment(
        horizontal="center"
    )



    ws.merge_cells("A2:H2")


    ws["A2"] = (

        "Generated Date : "

        +

        datetime.now().strftime(
            "%Y-%m-%d %H:%M"
        )

    )



    # ==========================================================================
    # HEADER
    # ==========================================================================


    headers = [

        "No",

        "Product Name",

        "SKU",

        "Category",

        "Purchase Cost",

        "Markup %",

        "Selling Price",

        "Profit"

    ]



    row = 4



    for col, header in enumerate(
        headers,
        start=1
    ):


        cell = ws.cell(
            row=row,
            column=col
        )


        cell.value = header


        cell.font = Font(
            bold=True
        )


        cell.alignment = Alignment(
            horizontal="center"
        )


        cell.border = thin_border



    # ==========================================================================
    # DATA
    # ==========================================================================


    total_profit = 0


    for index, product in enumerate(
        products,
        start=1
    ):


        cost = float(

            product.get(
                "purchase_price",
                0
            )
            or 0

        )


        selling = float(

            product.get(
                "selling_price",
                0
            )
            or 0

        )


        profit = selling - cost


        total_profit += profit



        data = [

            index,

            product.get(
                "name",
                ""
            ),

            product.get(
                "sku",
                ""
            ),

            product.get(
                "category",
                ""
            ),

            cost,

            product.get(
                "markup_percent",
                0
            ),

            selling,

            profit

        ]



        row += 1



        for col, value in enumerate(
            data,
            start=1
        ):


            cell = ws.cell(

                row=row,

                column=col

            )


            cell.value = value


            cell.border = thin_border



            if col in [5,7,8]:

                cell.number_format = (
                    '#,##0.00'
                )



    # ==========================================================================
    # SUMMARY
    # ==========================================================================


    row += 2



    ws.cell(
        row=row,
        column=1
    ).value = "Total Products"



    ws.cell(
        row=row,
        column=2
    ).value = len(products)



    ws.cell(
        row=row + 1,
        column=1
    ).value = "Total Profit"



    ws.cell(
        row=row + 1,
        column=2
    ).value = total_profit



    ws.cell(
        row=row + 1,
        column=2
    ).number_format = (
        '#,##0.00'
    )



    # ==========================================================================
    # AUTO WIDTH
    # ==========================================================================


    for column in ws.columns:


        max_length = 0


        column_letter = get_column_letter(
            column[0].column
        )


        for cell in column:


            try:

                if len(
                    str(cell.value)
                ) > max_length:

                    max_length = len(
                        str(cell.value)
                    )

            except:

                pass



        ws.column_dimensions[
            column_letter
        ].width = max_length + 3



    # ==========================================================================
    # RETURN EXCEL FILE
    # ==========================================================================


    output = BytesIO()


    wb.save(
        output
    )


    output.seek(
        0
    )


    return output
