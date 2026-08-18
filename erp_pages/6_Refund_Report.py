# ==============================================================================
# ERP ENTERPRISE REFUND REPORT v5.1
#
# TAX-AWARE + DATE RANGE + REJECTED SEPARATION
#
# IMPORTANT BUSINESS RULE
#
# ACTUAL REFUND TOTAL:
#     COMPLETED + APPROVED
#
# REJECTED:
#     NEVER INCLUDED IN REFUND TOTAL
#     SHOWN SEPARATELY AS REJECTED TOTAL
#
# PENDING:
#     NEVER INCLUDED IN REFUND TOTAL
#     SHOWN SEPARATELY AS PENDING TOTAL
#
# DATE:
#     FROM DATE -> TO DATE
#
# EXPORT:
#     PDF
#     EXCEL
#     CSV
#     HTML
#
# DATABASE VIEWS:
#     refund_report_view
#     refund_detail_view
#
# V4 AMOUNTS:
#     refund_net_amount
#     refund_tax_amount
#     refund_total_amount
# ==============================================================================


# ==============================================================================
# IMPORTS
# ==============================================================================

import io

from datetime import date, timedelta

import pandas as pd
import streamlit as st

from database import db
from auth import require_login

from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
)

from reportlab.lib import colors

from reportlab.lib.styles import (
    getSampleStyleSheet,
    ParagraphStyle,
)

from reportlab.lib.pagesizes import A4

from reportlab.lib.enums import (
    TA_LEFT,
    TA_CENTER,
)


# ==============================================================================
# PAGE CONFIG
# ==============================================================================

st.set_page_config(
    page_title="Refund Report",
    page_icon="↩️",
    layout="wide",
)


# ==============================================================================
# AUTH
# ==============================================================================

user = require_login()


# ==============================================================================
# TITLE
# ==============================================================================

st.title("↩️ Refund Report")

st.caption(
    "ERP Enterprise Refund Report • "
    "Date Range • Tax Aware • PDF • Excel • CSV • HTML"
)


# ==============================================================================
# BUSINESS STATUS RULE
# ==============================================================================

# ------------------------------------------------------------------------------
# ACTUAL REFUND
#
# Only these statuses are treated as real refund financial totals.
# ------------------------------------------------------------------------------

ACTUAL_REFUND_STATUSES = [
    "COMPLETED",
    "APPROVED",
]


# ------------------------------------------------------------------------------
# REJECTED
# ------------------------------------------------------------------------------

REJECTED_STATUS = "REJECTED"


# ------------------------------------------------------------------------------
# PENDING
# ------------------------------------------------------------------------------

PENDING_STATUS = "PENDING"


# ==============================================================================
# COLUMN DEFINITIONS
# ==============================================================================

NUMERIC_COLUMNS = [
    "quantity",
    "unit_price",
    "item_total",
    "refund_amount",
    "refund_net_amount",
    "refund_tax_amount",
    "refund_total_amount",
]


TEXT_COLUMNS = [
    "invoice_no",
    "cashier_name",
    "warehouse_name",
    "product_name",
    "processed_by",
    "reason",
]


# ==============================================================================
# SAFE HELPERS
# ==============================================================================

def safe_float(value):
    """
    Safely convert a value to float.
    """

    try:

        if value is None:
            return 0.0

        if pd.isna(value):
            return 0.0

        return float(value)

    except Exception:

        return 0.0


def safe_text(value):
    """
    Safely convert a value to string.
    """

    if value is None:
        return ""

    try:

        if pd.isna(value):
            return ""

    except Exception:
        pass

    return str(value)


def money(value):
    """
    MMK money formatter.
    """

    return (
        f"{safe_float(value):,.2f} MMK"
    )


# ==============================================================================
# NORMALIZE REPORT DATAFRAME
# ==============================================================================

def normalize_report_dataframe(df):

    if df is None:

        df = pd.DataFrame()

    df = df.copy()

    # ==========================================================================
    # NUMERIC COLUMNS
    # ==========================================================================

    for col in NUMERIC_COLUMNS:

        if col not in df.columns:

            df[col] = 0.0

        df[col] = pd.to_numeric(
            df[col],
            errors="coerce",
        ).fillna(0.0)

    # ==========================================================================
    # TEXT COLUMNS
    # ==========================================================================

    for col in TEXT_COLUMNS:

        if col not in df.columns:

            df[col] = ""

        df[col] = (
            df[col]
            .fillna("")
            .astype(str)
        )

    # ==========================================================================
    # ID COLUMNS
    # ==========================================================================

    for col in [
        "refund_id",
        "sale_id",
        "product_id",
    ]:

        if col not in df.columns:

            df[col] = ""

    # ==========================================================================
    # STATUS
    # ==========================================================================

    if "status" not in df.columns:

        df["status"] = "COMPLETED"

    df["status"] = (
        df["status"]
        .fillna("COMPLETED")
        .astype(str)
        .str.strip()
        .str.upper()
    )

    # ==========================================================================
    # REFUND DATE
    # ==========================================================================

    if "refund_date" not in df.columns:

        df["refund_date"] = pd.NaT

    df["refund_date"] = pd.to_datetime(
        df["refund_date"],
        errors="coerce",
    )

    # ==========================================================================
    # APPROVED AT
    # ==========================================================================

    if "approved_at" not in df.columns:

        df["approved_at"] = ""

    # ==========================================================================
    # V4 REPORT AMOUNTS
    # ==========================================================================

    df["report_net"] = (
        df["refund_net_amount"]
        .astype(float)
    )

    df["report_tax"] = (
        df["refund_tax_amount"]
        .astype(float)
    )

    df["report_total"] = (
        df["refund_total_amount"]
        .astype(float)
    )

    # ==========================================================================
    # LEGACY FALLBACK
    #
    # Old refund rows:
    #
    #     Net   = item_total
    #     Tax   = 0
    #     Total = item_total
    #
    # New V4 rows are not affected.
    # ==========================================================================

    legacy_total_mask = (
        (df["report_total"] == 0)
        &
        (df["item_total"] != 0)
    )

    df.loc[
        legacy_total_mask,
        "report_total",
    ] = df.loc[
        legacy_total_mask,
        "item_total",
    ]

    legacy_net_mask = (
        (df["report_net"] == 0)
        &
        (df["item_total"] != 0)
    )

    df.loc[
        legacy_net_mask,
        "report_net",
    ] = df.loc[
        legacy_net_mask,
        "item_total",
    ]

    return df


# ==============================================================================
# LOAD REFUND REPORT
# ==============================================================================

@st.cache_data(ttl=60)
def load_refund_report():

    response = (
        db()
        .table("refund_report_view")
        .select("*")
        .order(
            "refund_date",
            desc=True,
        )
        .execute()
    )

    return pd.DataFrame(
        response.data or []
    )


# ==============================================================================
# LOAD DATA
# ==============================================================================

try:

    df = load_refund_report()

except Exception as e:

    st.error(
        "Unable to load refund report."
    )

    st.exception(e)

    st.stop()


df = normalize_report_dataframe(df)


# ==============================================================================
# EMPTY DATA
# ==============================================================================

if df.empty:

    st.info(
        "No refund records found."
    )

    if st.button("🔄 Refresh"):

        st.cache_data.clear()
        st.rerun()

    st.stop()


# ==============================================================================
# SESSION DATE DEFAULT
# ==============================================================================

valid_dates = (
    df["refund_date"]
    .dropna()
)


if valid_dates.empty:

    default_from_date = date.today()

    default_to_date = date.today()

else:

    default_from_date = (
        valid_dates
        .min()
        .date()
    )

    default_to_date = date.today()

    if default_to_date < default_from_date:

        default_to_date = (
            default_from_date
        )


if "refund_from_date" not in st.session_state:

    st.session_state.refund_from_date = (
        default_from_date
    )


if "refund_to_date" not in st.session_state:

    st.session_state.refund_to_date = (
        default_to_date
    )


# ==============================================================================
# SIDEBAR
# ==============================================================================

st.sidebar.header(
    "🔍 Report Filter"
)


# ==============================================================================
# DATE RANGE
# ==============================================================================

st.sidebar.subheader(
    "📅 Report Period"
)


# ==============================================================================
# QUICK DATE BUTTONS
# ==============================================================================

q1, q2 = st.sidebar.columns(2)


with q1:

    today_clicked = st.button(
        "Today",
        use_container_width=True,
    )


with q2:

    yesterday_clicked = st.button(
        "Yesterday",
        use_container_width=True,
    )


q3, q4 = st.sidebar.columns(2)


with q3:

    month_clicked = st.button(
        "This Month",
        use_container_width=True,
    )


with q4:

    last_month_clicked = st.button(
        "Last Month",
        use_container_width=True,
    )


# ==============================================================================
# QUICK DATE LOGIC
# ==============================================================================

today = date.today()


if today_clicked:

    st.session_state.refund_from_date = (
        today
    )

    st.session_state.refund_to_date = (
        today
    )


elif yesterday_clicked:

    yesterday = (
        today
        - timedelta(days=1)
    )

    st.session_state.refund_from_date = (
        yesterday
    )

    st.session_state.refund_to_date = (
        yesterday
    )


elif month_clicked:

    first_day = today.replace(
        day=1
    )

    st.session_state.refund_from_date = (
        first_day
    )

    st.session_state.refund_to_date = (
        today
    )


elif last_month_clicked:

    first_this_month = (
        today.replace(day=1)
    )

    last_previous_month = (
        first_this_month
        - timedelta(days=1)
    )

    first_previous_month = (
        last_previous_month
        .replace(day=1)
    )

    st.session_state.refund_from_date = (
        first_previous_month
    )

    st.session_state.refund_to_date = (
        last_previous_month
    )


# ==============================================================================
# FROM DATE
# ==============================================================================

from_date = st.sidebar.date_input(
    "From Date",
    value=st.session_state.refund_from_date,
)


# ==============================================================================
# TO DATE
# ==============================================================================

to_date = st.sidebar.date_input(
    "To Date",
    value=st.session_state.refund_to_date,
)


# ==============================================================================
# SAVE DATE SESSION
# ==============================================================================

st.session_state.refund_from_date = (
    from_date
)

st.session_state.refund_to_date = (
    to_date
)


# ==============================================================================
# DATE VALIDATION
# ==============================================================================

if from_date > to_date:

    st.sidebar.error(
        "From Date cannot be later than To Date."
    )

    st.stop()


# ==============================================================================
# INVOICE SEARCH
# ==============================================================================

invoice_search = st.sidebar.text_input(
    "Invoice No",
    placeholder="Search invoice...",
)


# ==============================================================================
# CASHIER FILTER
# ==============================================================================

cashiers = sorted(
    [
        safe_text(x)
        for x in df["cashier_name"].unique()
        if safe_text(x).strip()
    ]
)


cashier_filter = st.sidebar.multiselect(
    "Cashier",
    cashiers,
)


# ==============================================================================
# WAREHOUSE FILTER
# ==============================================================================

warehouses = sorted(
    [
        safe_text(x)
        for x in df["warehouse_name"].unique()
        if safe_text(x).strip()
    ]
)


warehouse_filter = st.sidebar.multiselect(
    "Warehouse",
    warehouses,
)


# ==============================================================================
# STATUS FILTER
# ==============================================================================

statuses = sorted(
    [
        safe_text(x)
        for x in df["status"].unique()
        if safe_text(x).strip()
    ]
)


status_filter = st.sidebar.multiselect(
    "Status",
    statuses,
)


# ==============================================================================
# REFRESH
# ==============================================================================

st.sidebar.divider()


if st.sidebar.button(
    "🔄 Refresh Data",
    use_container_width=True,
):

    st.cache_data.clear()

    st.rerun()


# ==============================================================================
# APPLY DATE FILTER
# ==============================================================================

filtered = df.copy()


filtered = filtered[
    (
        filtered["refund_date"].dt.date
        >= from_date
    )
    &
    (
        filtered["refund_date"].dt.date
        <= to_date
    )
]


# ==============================================================================
# APPLY INVOICE FILTER
# ==============================================================================

if invoice_search:

    filtered = filtered[
        filtered[
            "invoice_no"
        ].str.contains(
            invoice_search,
            case=False,
            na=False,
        )
    ]


# ==============================================================================
# APPLY CASHIER FILTER
# ==============================================================================

if cashier_filter:

    filtered = filtered[
        filtered[
            "cashier_name"
        ].isin(
            cashier_filter
        )
    ]


# ==============================================================================
# APPLY WAREHOUSE FILTER
# ==============================================================================

if warehouse_filter:

    filtered = filtered[
        filtered[
            "warehouse_name"
        ].isin(
            warehouse_filter
        )
    ]


# ==============================================================================
# APPLY STATUS FILTER
# ==============================================================================

if status_filter:

    filtered = filtered[
        filtered[
            "status"
        ].isin(
            status_filter
        )
    ]


# ==============================================================================
# REPORT PERIOD
# ==============================================================================

st.info(
    f"📅 Report Period: "
    f"**{from_date.strftime('%Y-%m-%d')}** "
    f"→ "
    f"**{to_date.strftime('%Y-%m-%d')}**"
)


# ==============================================================================
# FINANCIAL STATUS MASKS
# ==============================================================================

# ------------------------------------------------------------------------------
# ACTUAL REFUND
#
# COMPLETED + APPROVED only
# ------------------------------------------------------------------------------

actual_refund_mask = (
    filtered["status"]
    .isin(
        ACTUAL_REFUND_STATUSES
    )
)


# ------------------------------------------------------------------------------
# REJECTED
# ------------------------------------------------------------------------------

rejected_mask = (
    filtered["status"]
    == REJECTED_STATUS
)


# ------------------------------------------------------------------------------
# PENDING
# ------------------------------------------------------------------------------

pending_mask = (
    filtered["status"]
    == PENDING_STATUS
)


# ==============================================================================
# COUNTS
# ==============================================================================

total_refunds = (
    filtered["refund_id"]
    .nunique()
)


completed_count = (
    filtered["status"]
    == "COMPLETED"
).sum()


approved_count = (
    filtered["status"]
    == "APPROVED"
).sum()


pending_count = (
    pending_mask
).sum()


rejected_count = (
    rejected_mask
).sum()


actual_refund_count = (
    filtered.loc[
        actual_refund_mask,
        "refund_id",
    ]
    .nunique()
)


# ==============================================================================
# ACTUAL REFUND FINANCIAL TOTAL
#
# IMPORTANT:
# REJECTED IS NOT INCLUDED
# PENDING IS NOT INCLUDED
# ==============================================================================

total_net = (
    filtered.loc[
        actual_refund_mask,
        "report_net",
    ]
    .sum()
)


total_tax = (
    filtered.loc[
        actual_refund_mask,
        "report_tax",
    ]
    .sum()
)


total_refund = (
    filtered.loc[
        actual_refund_mask,
        "report_total",
    ]
    .sum()
)


# ==============================================================================
# REJECTED TOTAL
#
# COMPLETELY SEPARATE
# ==============================================================================

rejected_net = (
    filtered.loc[
        rejected_mask,
        "report_net",
    ]
    .sum()
)


rejected_tax = (
    filtered.loc[
        rejected_mask,
        "report_tax",
    ]
    .sum()
)


rejected_total = (
    filtered.loc[
        rejected_mask,
        "report_total",
    ]
    .sum()
)


# ==============================================================================
# PENDING TOTAL
#
# COMPLETELY SEPARATE
# ==============================================================================

pending_net = (
    filtered.loc[
        pending_mask,
        "report_net",
    ]
    .sum()
)


pending_tax = (
    filtered.loc[
        pending_mask,
        "report_tax",
    ]
    .sum()
)


pending_total = (
    filtered.loc[
        pending_mask,
        "report_total",
    ]
    .sum()
)


# ==============================================================================
# SUMMARY
# ==============================================================================

st.subheader(
    "📊 Refund Summary"
)


c1, c2, c3, c4, c5 = st.columns(5)


# ------------------------------------------------------------------------------
# REFUND COUNT
# ------------------------------------------------------------------------------

with c1:

    st.metric(
        "Refund Records",
        f"{total_refunds:,}",
    )


# ------------------------------------------------------------------------------
# COMPLETED
# ------------------------------------------------------------------------------

with c2:

    st.metric(
        "Completed",
        f"{completed_count:,}",
    )


# ------------------------------------------------------------------------------
# PENDING
# ------------------------------------------------------------------------------

with c3:

    st.metric(
        "Pending",
        f"{pending_count:,}",
    )


# ------------------------------------------------------------------------------
# REJECTED
# ------------------------------------------------------------------------------

with c4:

    st.metric(
        "Rejected",
        f"{rejected_count:,}",
    )


# ------------------------------------------------------------------------------
# ACTUAL REFUND TOTAL
# ------------------------------------------------------------------------------

with c5:

    st.metric(
        "Refund Total",
        money(total_refund),
    )


# ==============================================================================
# FINANCIAL SUMMARY
# ==============================================================================

st.divider()

st.subheader(
    "💰 Financial Summary"
)


f1, f2, f3, f4 = st.columns(4)


with f1:

    st.metric(
        "Refund Net",
        money(total_net),
    )


with f2:

    st.metric(
        "Refund Tax",
        money(total_tax),
    )


with f3:

    st.metric(
        "Refund Total",
        money(total_refund),
    )


with f4:

    st.metric(
        "Rejected Total",
        money(rejected_total),
    )


# ==============================================================================
# PENDING FINANCIAL SUMMARY
# ==============================================================================

if pending_total != 0:

    st.warning(
        f"⏳ Pending Refund Total: "
        f"**{money(pending_total)}**"
    )


# ==============================================================================
# REJECTED FINANCIAL SUMMARY
# ==============================================================================

if rejected_total != 0:

    st.error(
        f"❌ Rejected Refund Total: "
        f"**{money(rejected_total)}**"
    )


# ==============================================================================
# REFUND REGISTER
# ==============================================================================

st.divider()

st.subheader(
    "📋 Refund Register"
)


if filtered.empty:

    st.warning(
        "No refund records match the selected filters."
    )

else:

    register_df = pd.DataFrame(
        {
            "Refund ID":
                filtered[
                    "refund_id"
                ],

            "Invoice":
                filtered[
                    "invoice_no"
                ],

            "Refund Date":
                filtered[
                    "refund_date"
                ].dt.strftime(
                    "%Y-%m-%d %H:%M"
                ),

            "Status":
                filtered[
                    "status"
                ],

            "Product":
                filtered[
                    "product_name"
                ],

            "Qty":
                filtered[
                    "quantity"
                ],

            "Refund Net":
                filtered[
                    "report_net"
                ],

            "Refund Tax":
                filtered[
                    "report_tax"
                ],

            "Refund Total":
                filtered[
                    "report_total"
                ],

            "Cashier":
                filtered[
                    "cashier_name"
                ],

            "Warehouse":
                filtered[
                    "warehouse_name"
                ],
        }
    )


    st.dataframe(
        register_df,
        use_container_width=True,
        hide_index=True,

        column_config={

            "Refund ID":
                st.column_config.TextColumn(
                    "Refund ID",
                ),

            "Refund Date":
                st.column_config.TextColumn(
                    "Refund Date",
                ),

            "Qty":
                st.column_config.NumberColumn(
                    "Qty",
                    format="%.2f",
                ),

            "Refund Net":
                st.column_config.NumberColumn(
                    "Refund Net",
                    format="%,.2f MMK",
                ),

            "Refund Tax":
                st.column_config.NumberColumn(
                    "Refund Tax",
                    format="%,.2f MMK",
                ),

            "Refund Total":
                st.column_config.NumberColumn(
                    "Refund Total",
                    format="%,.2f MMK",
                ),
        },
    )


# ==============================================================================
# PDF HEADER / FOOTER
# ==============================================================================

def pdf_header_footer(
    canvas,
    doc,
):

    canvas.saveState()

    width, height = A4

    # --------------------------------------------------------------------------
    # HEADER
    # --------------------------------------------------------------------------

    canvas.setFont(
        "Helvetica-Bold",
        9,
    )

    canvas.drawString(
        40,
        height - 28,
        "ERP ENTERPRISE",
    )

    canvas.setFont(
        "Helvetica",
        8,
    )

    canvas.drawRightString(
        width - 40,
        height - 28,
        "REFUND REPORT",
    )

    # --------------------------------------------------------------------------
    # FOOTER
    # --------------------------------------------------------------------------

    canvas.setFont(
        "Helvetica",
        8,
    )

    canvas.drawString(
        40,
        22,
        "ERP Refund Report",
    )

    canvas.drawRightString(
        width - 40,
        22,
        f"Page {doc.page}",
    )

    canvas.restoreState()


# ==============================================================================
# CREATE PDF REPORT
# ==============================================================================

def create_refund_report_pdf(
    report_df,
    report_from,
    report_to,
):

    buffer = io.BytesIO()

    # ==========================================================================
    # ACTUAL REFUND DATA
    # ==========================================================================

    actual_mask = (
        report_df["status"]
        .isin(
            ACTUAL_REFUND_STATUSES
        )
    )

    rejected_mask_pdf = (
        report_df["status"]
        == REJECTED_STATUS
    )

    pending_mask_pdf = (
        report_df["status"]
        == PENDING_STATUS
    )

    # ==========================================================================
    # TOTALS
    # ==========================================================================

    actual_net = (
        report_df.loc[
            actual_mask,
            "report_net",
        ]
        .sum()
    )

    actual_tax = (
        report_df.loc[
            actual_mask,
            "report_tax",
        ]
        .sum()
    )

    actual_total = (
        report_df.loc[
            actual_mask,
            "report_total",
        ]
        .sum()
    )

    rejected_total_pdf = (
        report_df.loc[
            rejected_mask_pdf,
            "report_total",
        ]
        .sum()
    )

    pending_total_pdf = (
        report_df.loc[
            pending_mask_pdf,
            "report_total",
        ]
        .sum()
    )

    refund_count_pdf = (
        report_df.loc[
            actual_mask,
            "refund_id",
        ]
        .nunique()
    )

    rejected_count_pdf = (
        report_df.loc[
            rejected_mask_pdf,
            "refund_id",
        ]
        .nunique()
    )

    pending_count_pdf = (
        report_df.loc[
            pending_mask_pdf,
            "refund_id",
        ]
        .nunique()
    )

    # ==========================================================================
    # DOCUMENT
    # ==========================================================================

    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,

        # ----------------------------------------------------------------------
        # PAGE MARGINS
        # ----------------------------------------------------------------------

        leftMargin=40,
        rightMargin=40,
        topMargin=48,
        bottomMargin=40,
    )

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "RefundTitle",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=18,
        leading=22,
        alignment=TA_CENTER,
        spaceAfter=6,
    )

    subtitle_style = ParagraphStyle(
        "RefundSubtitle",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=9,
        leading=12,
        alignment=TA_CENTER,
        spaceAfter=14,
    )

    small_style = ParagraphStyle(
        "RefundSmall",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        leading=10,
        alignment=TA_LEFT,
    )

    content = []

    # ==========================================================================
    # TITLE
    # ==========================================================================

    content.append(
        Paragraph(
            "REFUND REPORT",
            title_style,
        )
    )

    content.append(
        Paragraph(
            (
                f"Report Period: "
                f"<b>{report_from.strftime('%Y-%m-%d')}</b>"
                f" to "
                f"<b>{report_to.strftime('%Y-%m-%d')}</b>"
            ),
            subtitle_style,
        )
    )

    # ==========================================================================
    # FINANCIAL SUMMARY TABLE
    # ==========================================================================

    summary_data = [
        [
            "Actual Refunds",
            "Refund Net",
            "Refund Tax",
            "Refund Total",
        ],

        [
            f"{refund_count_pdf:,}",
            f"{actual_net:,.2f}",
            f"{actual_tax:,.2f}",
            f"{actual_total:,.2f}",
        ],

        [
            "Rejected",
            "Rejected Total",
            "Pending",
            "Pending Total",
        ],

        [
            f"{rejected_count_pdf:,}",
            f"{rejected_total_pdf:,.2f}",
            f"{pending_count_pdf:,}",
            f"{pending_total_pdf:,.2f}",
        ],
    ]

    summary_table = Table(
        summary_data,
        colWidths=[
            125,
            125,
            125,
            125,
        ],
    )

    summary_table.setStyle(
        TableStyle(
            [
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    colors.HexColor(
                        "#343A40"
                    ),
                ),

                (
                    "TEXTCOLOR",
                    (0, 0),
                    (-1, 0),
                    colors.white,
                ),

                (
                    "BACKGROUND",
                    (0, 2),
                    (-1, 2),
                    colors.HexColor(
                        "#E9ECEF"
                    ),
                ),

                (
                    "FONTNAME",
                    (0, 0),
                    (-1, 0),
                    "Helvetica-Bold",
                ),

                (
                    "FONTNAME",
                    (0, 2),
                    (-1, 2),
                    "Helvetica-Bold",
                ),

                (
                    "FONTNAME",
                    (0, 1),
                    (-1, 1),
                    "Helvetica-Bold",
                ),

                (
                    "FONTNAME",
                    (0, 3),
                    (-1, 3),
                    "Helvetica-Bold",
                ),

                (
                    "FONTSIZE",
                    (0, 0),
                    (-1, -1),
                    8,
                ),

                (
                    "ALIGN",
                    (0, 0),
                    (-1, -1),
                    "CENTER",
                ),

                (
                    "VALIGN",
                    (0, 0),
                    (-1, -1),
                    "MIDDLE",
                ),

                (
                    "GRID",
                    (0, 0),
                    (-1, -1),
                    0.4,
                    colors.HexColor(
                        "#BFC3C7"
                    ),
                ),

                # ----------------------------------------------------------------
                # TABLE MARGIN / PADDING
                # ----------------------------------------------------------------

                (
                    "TOPPADDING",
                    (0, 0),
                    (-1, -1),
                    7,
                ),

                (
                    "BOTTOMPADDING",
                    (0, 0),
                    (-1, -1),
                    7,
                ),

                (
                    "LEFTPADDING",
                    (0, 0),
                    (-1, -1),
                    6,
                ),

                (
                    "RIGHTPADDING",
                    (0, 0),
                    (-1, -1),
                    6,
                ),
            ]
        )
    )

    content.append(
        summary_table
    )

    content.append(
        Spacer(1, 16)
    )

    # ==========================================================================
    # DETAIL TABLE
    # ==========================================================================

    table_data = [
        [
            "ID",
            "Invoice",
            "Date",
            "Status",
            "Product",
            "Qty",
            "Net",
            "Tax",
            "Total",
        ]
    ]

    for _, row in report_df.iterrows():

        refund_date = row[
            "refund_date"
        ]

        if pd.notna(
            refund_date
        ):

            date_text = (
                refund_date.strftime(
                    "%Y-%m-%d"
                )
            )

        else:

            date_text = ""

        table_data.append(
            [
                safe_text(
                    row["refund_id"]
                ),

                safe_text(
                    row["invoice_no"]
                ),

                date_text,

                safe_text(
                    row["status"]
                ),

                safe_text(
                    row["product_name"]
                ),

                f"{safe_float(row['quantity']):,.2f}",

                f"{safe_float(row['report_net']):,.2f}",

                f"{safe_float(row['report_tax']):,.2f}",

                f"{safe_float(row['report_total']):,.2f}",
            ]
        )

    # ==========================================================================
    # ACTUAL REFUND TOTAL ROW
    #
    # IMPORTANT:
    # REJECTED / PENDING ARE NOT INCLUDED
    # ==========================================================================

    table_data.append(
        [
            "",
            "",
            "",
            "",
            "REFUND TOTAL",
            "",
            f"{actual_net:,.2f}",
            f"{actual_tax:,.2f}",
            f"{actual_total:,.2f}",
        ]
    )

    # ==========================================================================
    # REJECTED TOTAL ROW
    # ==========================================================================

    table_data.append(
        [
            "",
            "",
            "",
            "",
            "REJECTED TOTAL",
            "",
            "",
            "",
            f"{rejected_total_pdf:,.2f}",
        ]
    )

    # ==========================================================================
    # PENDING TOTAL ROW
    # ==========================================================================

    table_data.append(
        [
            "",
            "",
            "",
            "",
            "PENDING TOTAL",
            "",
            "",
            "",
            f"{pending_total_pdf:,.2f}",
        ]
    )

    detail_table = Table(
        table_data,
        repeatRows=1,

        colWidths=[
            35,
            60,
            58,
            55,
            105,
            35,
            58,
            48,
            61,
        ],
    )

    detail_table.setStyle(
        TableStyle(
            [
                # ----------------------------------------------------------------
                # HEADER
                # ----------------------------------------------------------------

                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    colors.HexColor(
                        "#343A40"
                    ),
                ),

                (
                    "TEXTCOLOR",
                    (0, 0),
                    (-1, 0),
                    colors.white,
                ),

                (
                    "FONTNAME",
                    (0, 0),
                    (-1, 0),
                    "Helvetica-Bold",
                ),

                (
                    "FONTSIZE",
                    (0, 0),
                    (-1, 0),
                    7,
                ),

                # ----------------------------------------------------------------
                # BODY
                # ----------------------------------------------------------------

                (
                    "FONTNAME",
                    (0, 1),
                    (-1, -4),
                    "Helvetica",
                ),

                (
                    "FONTSIZE",
                    (0, 1),
                    (-1, -1),
                    6.8,
                ),

                # ----------------------------------------------------------------
                # REFUND TOTAL
                # ----------------------------------------------------------------

                (
                    "BACKGROUND",
                    (0, -3),
                    (-1, -3),
                    colors.HexColor(
                        "#E9ECEF"
                    ),
                ),

                (
                    "FONTNAME",
                    (0, -3),
                    (-1, -3),
                    "Helvetica-Bold",
                ),

                # ----------------------------------------------------------------
                # REJECTED TOTAL
                # ----------------------------------------------------------------

                (
                    "BACKGROUND",
                    (0, -2),
                    (-1, -2),
                    colors.HexColor(
                        "#F8D7DA"
                    ),
                ),

                (
                    "FONTNAME",
                    (0, -2),
                    (-1, -2),
                    "Helvetica-Bold",
                ),

                # ----------------------------------------------------------------
                # PENDING TOTAL
                # ----------------------------------------------------------------

                (
                    "BACKGROUND",
                    (0, -1),
                    (-1, -1),
                    colors.HexColor(
                        "#FFF3CD"
                    ),
                ),

                (
                    "FONTNAME",
                    (0, -1),
                    (-1, -1),
                    "Helvetica-Bold",
                ),

                # ----------------------------------------------------------------
                # GRID
                # ----------------------------------------------------------------

                (
                    "GRID",
                    (0, 0),
                    (-1, -1),
                    0.35,
                    colors.HexColor(
                        "#BFC3C7"
                    ),
                ),

                # ----------------------------------------------------------------
                # ALIGN
                # ----------------------------------------------------------------

                (
                    "ALIGN",
                    (0, 0),
                    (-1, -1),
                    "LEFT",
                ),

                (
                    "ALIGN",
                    (5, 1),
                    (-1, -1),
                    "RIGHT",
                ),

                (
                    "ALIGN",
                    (5, 0),
                    (-1, 0),
                    "CENTER",
                ),

                (
                    "VALIGN",
                    (0, 0),
                    (-1, -1),
                    "MIDDLE",
                ),

                # ----------------------------------------------------------------
                # PADDING
                # ----------------------------------------------------------------

                (
                    "TOPPADDING",
                    (0, 0),
                    (-1, -1),
                    5,
                ),

                (
                    "BOTTOMPADDING",
                    (0, 0),
                    (-1, -1),
                    5,
                ),

                (
                    "LEFTPADDING",
                    (0, 0),
                    (-1, -1),
                    5,
                ),

                (
                    "RIGHTPADDING",
                    (0, 0),
                    (-1, -1),
                    5,
                ),
            ]
        )
    )

    content.append(
        detail_table
    )

    content.append(
        Spacer(1, 14)
    )

    # ==========================================================================
    # FINAL NOTE
    # ==========================================================================

    content.append(
        Paragraph(
            (
                "<b>Accounting Rule:</b> "
                "Refund Total includes only COMPLETED and APPROVED refunds. "
                "Rejected and Pending amounts are reported separately."
            ),
            small_style,
        )
    )

    # ==========================================================================
    # BUILD PDF
    # ==========================================================================

    doc.build(
        content,
        onFirstPage=pdf_header_footer,
        onLaterPages=pdf_header_footer,
    )

    buffer.seek(0)

    return buffer.getvalue()


# ==============================================================================
# CREATE HTML REPORT
# ==============================================================================

def create_html_report(
    report_df,
    report_from,
    report_to,
):

    actual_mask = (
        report_df["status"]
        .isin(
            ACTUAL_REFUND_STATUSES
        )
    )

    rejected_mask_html = (
        report_df["status"]
        == REJECTED_STATUS
    )

    pending_mask_html = (
        report_df["status"]
        == PENDING_STATUS
    )

    actual_net_html = (
        report_df.loc[
            actual_mask,
            "report_net",
        ]
        .sum()
    )

    actual_tax_html = (
        report_df.loc[
            actual_mask,
            "report_tax",
        ]
        .sum()
    )

    actual_total_html = (
        report_df.loc[
            actual_mask,
            "report_total",
        ]
        .sum()
    )

    rejected_total_html = (
        report_df.loc[
            rejected_mask_html,
            "report_total",
        ]
        .sum()
    )

    pending_total_html = (
        report_df.loc[
            pending_mask_html,
            "report_total",
        ]
        .sum()
    )

    actual_count_html = (
        report_df.loc[
            actual_mask,
            "refund_id",
        ]
        .nunique()
    )

    rejected_count_html = (
        report_df.loc[
            rejected_mask_html,
            "refund_id",
        ]
        .nunique()
    )

    pending_count_html = (
        report_df.loc[
            pending_mask_html,
            "refund_id",
        ]
        .nunique()
    )

    html_df = report_df.copy()

    html_df["Refund Date"] = (
        html_df["refund_date"]
        .dt.strftime(
            "%Y-%m-%d %H:%M"
        )
    )

    display_df = pd.DataFrame(
        {
            "Refund ID":
                html_df["refund_id"],

            "Invoice":
                html_df["invoice_no"],

            "Refund Date":
                html_df["Refund Date"],

            "Status":
                html_df["status"],

            "Product":
                html_df["product_name"],

            "Qty":
                html_df["quantity"],

            "Refund Net":
                html_df["report_net"],

            "Refund Tax":
                html_df["report_tax"],

            "Refund Total":
                html_df["report_total"],

            "Cashier":
                html_df["cashier_name"],

            "Warehouse":
                html_df["warehouse_name"],
        }
    )

    table_html = display_df.to_html(
        index=False,
        classes="refund-table",
        border=0,
    )

    return f"""
<!DOCTYPE html>

<html>

<head>

<meta charset="utf-8">

<title>Refund Report</title>

<style>

body {{
    font-family: Arial, sans-serif;
    margin: 40px;
    color: #222;
}}

h1 {{
    text-align: center;
    margin-bottom: 5px;
}}

.period {{
    text-align: center;
    color: #666;
    margin-bottom: 25px;
}}

.summary {{
    width: 100%;
    border-collapse: collapse;
    margin-bottom: 25px;
}}

.summary td {{
    border: 1px solid #ddd;
    padding: 12px;
    text-align: center;
}}

.summary-title {{
    font-weight: bold;
    background: #343a40;
    color: white;
}}

.reject-title {{
    font-weight: bold;
    background: #f8d7da;
}}

.pending-title {{
    font-weight: bold;
    background: #fff3cd;
}}

.refund-table {{
    width: 100%;
    border-collapse: collapse;
    margin-top: 20px;
}}

.refund-table th {{
    background: #343a40;
    color: white;
    padding: 9px;
    text-align: left;
}}

.refund-table td {{
    border: 1px solid #ddd;
    padding: 8px;
}}

.refund-table tr:nth-child(even) {{
    background: #f8f9fa;
}}

.footer {{
    margin-top: 25px;
    font-weight: bold;
}}

</style>

</head>

<body>

<h1>REFUND REPORT</h1>

<div class="period">

Report Period:
<b>{report_from.strftime("%Y-%m-%d")}</b>
to
<b>{report_to.strftime("%Y-%m-%d")}</b>

</div>


<table class="summary">

<tr>

<td class="summary-title">
Actual Refunds
</td>

<td class="summary-title">
Refund Net
</td>

<td class="summary-title">
Refund Tax
</td>

<td class="summary-title">
Refund Total
</td>

</tr>


<tr>

<td>
{actual_count_html:,}
</td>

<td>
{actual_net_html:,.2f} MMK
</td>

<td>
{actual_tax_html:,.2f} MMK
</td>

<td>
{actual_total_html:,.2f} MMK
</td>

</tr>


<tr>

<td class="reject-title">
Rejected
</td>

<td class="reject-title">
Rejected Total
</td>

<td class="pending-title">
Pending
</td>

<td class="pending-title">
Pending Total
</td>

</tr>


<tr>

<td>
{rejected_count_html:,}
</td>

<td>
{rejected_total_html:,.2f} MMK
</td>

<td>
{pending_count_html:,}
</td>

<td>
{pending_total_html:,.2f} MMK
</td>

</tr>

</table>


{table_html}


<div class="footer">

Actual Refund Total:
{actual_total_html:,.2f} MMK

<br><br>

Rejected Total:
{rejected_total_html:,.2f} MMK

<br><br>

Pending Total:
{pending_total_html:,.2f} MMK

</div>


</body>

</html>
"""


# ==============================================================================
# BUILD EXPORT DATAFRAME
# ==============================================================================

def build_export_dataframe(
    report_df,
):

    export_df = report_df.copy()

    # ==========================================================================
    # DATE
    # ==========================================================================

    if "refund_date" in export_df.columns:

        export_df["refund_date"] = (
            export_df[
                "refund_date"
            ]
            .dt.strftime(
                "%Y-%m-%d %H:%M:%S"
            )
        )

    # ==========================================================================
    # EXPORT COLUMNS
    # ==========================================================================

    export_columns = [
        "refund_id",
        "sale_id",
        "invoice_no",
        "refund_date",
        "status",
        "reason",
        "product_id",
        "product_name",
        "quantity",
        "unit_price",
        "item_total",
        "refund_net_amount",
        "refund_tax_amount",
        "refund_total_amount",
        "cashier_name",
        "processed_by",
        "approved_at",
        "warehouse_name",
    ]

    for col in export_columns:

        if col not in export_df.columns:

            export_df[col] = ""

    export_df = export_df[
        export_columns
    ].copy()

    # ==========================================================================
    # RENAME
    # ==========================================================================

    export_df = export_df.rename(
        columns={

            "refund_id":
                "Refund ID",

            "sale_id":
                "Sale ID",

            "invoice_no":
                "Invoice",

            "refund_date":
                "Refund Date",

            "status":
                "Status",

            "reason":
                "Reason",

            "product_id":
                "Product ID",

            "product_name":
                "Product",

            "quantity":
                "Quantity",

            "unit_price":
                "Unit Price",

            "item_total":
                "Item Price Total",

            "refund_net_amount":
                "Refund Net",

            "refund_tax_amount":
                "Refund Tax",

            "refund_total_amount":
                "Refund Total",

            "cashier_name":
                "Cashier",

            "processed_by":
                "Processed By",

            "approved_at":
                "Approved At",

            "warehouse_name":
                "Warehouse",
        }
    )

    return export_df


# ==============================================================================
# EXPORT SECTION
# ==============================================================================

st.divider()

st.subheader(
    "📥 Export Report"
)


if filtered.empty:

    st.info(
        "No data available for export."
    )

else:

    # ==========================================================================
    # EXPORT DATA
    # ==========================================================================

    export_df = build_export_dataframe(
        filtered
    )

    # ==========================================================================
    # PDF
    # ==========================================================================

    pdf_bytes = create_refund_report_pdf(
        filtered,
        from_date,
        to_date,
    )

    # ==========================================================================
    # EXCEL
    # ==========================================================================

    excel_buffer = io.BytesIO()

    with pd.ExcelWriter(
        excel_buffer,
        engine="openpyxl",
    ) as writer:

        export_df.to_excel(
            writer,
            index=False,
            sheet_name="Refund Report",
        )

        worksheet = writer.sheets[
            "Refund Report"
        ]

        # ======================================================================
        # FREEZE HEADER
        # ======================================================================

        worksheet.freeze_panes = "A2"

        # ======================================================================
        # AUTO FILTER
        # ======================================================================

        if worksheet.max_row >= 2:

            worksheet.auto_filter.ref = (
                worksheet.dimensions
            )

        # ======================================================================
        # COLUMN WIDTH
        # ======================================================================

        widths = {
            "A": 12,
            "B": 12,
            "C": 18,
            "D": 21,
            "E": 14,
            "F": 30,
            "G": 14,
            "H": 30,
            "I": 12,
            "J": 16,
            "K": 18,
            "L": 18,
            "M": 18,
            "N": 18,
            "O": 20,
            "P": 20,
            "Q": 20,
            "R": 20,
        }

        for letter, width in widths.items():

            worksheet.column_dimensions[
                letter
            ].width = width

        # ======================================================================
        # OPENPYXL FORMATTING
        # ======================================================================

        from openpyxl.styles import (
            Font,
            Alignment,
            PatternFill,
            Border,
            Side,
        )

        header_fill = PatternFill(
            "solid",
            fgColor="343A40",
        )

        header_font = Font(
            bold=True,
            color="FFFFFF",
        )

        thin_side = Side(
            style="thin",
            color="D9D9D9",
        )

        border = Border(
            left=thin_side,
            right=thin_side,
            top=thin_side,
            bottom=thin_side,
        )

        # ======================================================================
        # HEADER
        # ======================================================================

        for cell in worksheet[1]:

            cell.fill = header_fill

            cell.font = header_font

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

            cell.border = border

        # ======================================================================
        # BODY
        # ======================================================================

        for row in worksheet.iter_rows(
            min_row=2,
            max_row=worksheet.max_row,
        ):

            for cell in row:

                cell.border = border

                cell.alignment = Alignment(
                    vertical="center",
                )

        # ======================================================================
        # CURRENCY FORMAT
        # ======================================================================

        header_map = {
            cell.value:
                cell.column
            for cell in worksheet[1]
        }

        currency_columns = [
            "Unit Price",
            "Item Price Total",
            "Refund Net",
            "Refund Tax",
            "Refund Total",
        ]

        for column_name in currency_columns:

            if column_name not in header_map:

                continue

            column_number = (
                header_map[
                    column_name
                ]
            )

            for row in worksheet.iter_rows(
                min_row=2,
                max_row=worksheet.max_row,
                min_col=column_number,
                max_col=column_number,
            ):

                for cell in row:

                    cell.number_format = (
                        '#,##0.00'
                    )

    excel_buffer.seek(0)

    # ==========================================================================
    # CSV
    # ==========================================================================

    csv_bytes = (
        export_df
        .to_csv(
            index=False
        )
        .encode(
            "utf-8-sig"
        )
    )

    # ==========================================================================
    # HTML
    # ==========================================================================

    html_content = create_html_report(
        filtered,
        from_date,
        to_date,
    )

    # ==========================================================================
    # DOWNLOAD BUTTONS
    # ==========================================================================

    e1, e2, e3, e4 = st.columns(4)

    with e1:

        st.download_button(
            "📄 PDF",
            data=pdf_bytes,
            file_name=(
                f"refund_report_"
                f"{from_date}_to_{to_date}.pdf"
            ),
            mime="application/pdf",
            use_container_width=True,
        )

    with e2:

        st.download_button(
            "📊 Excel",
            data=(
                excel_buffer
                .getvalue()
            ),
            file_name=(
                f"refund_report_"
                f"{from_date}_to_{to_date}.xlsx"
            ),
            mime=(
                "application/vnd.openxmlformats-"
                "officedocument.spreadsheetml.sheet"
            ),
            use_container_width=True,
        )

    with e3:

        st.download_button(
            "📑 CSV",
            data=csv_bytes,
            file_name=(
                f"refund_report_"
                f"{from_date}_to_{to_date}.csv"
            ),
            mime="text/csv",
            use_container_width=True,
        )

    with e4:

        st.download_button(
            "🌐 HTML",
            data=html_content,
            file_name=(
                f"refund_report_"
                f"{from_date}_to_{to_date}.html"
            ),
            mime="text/html",
            use_container_width=True,
        )


# ==============================================================================
# SELECTED REFUND DETAIL
# ==============================================================================

st.divider()

st.subheader(
    "🔎 Refund Detail"
)


if filtered.empty:

    st.info(
        "No refund records available."
    )

else:

    # ==========================================================================
    # SELECTOR
    # ==========================================================================

    selector_items = []

    for _, row in filtered.iterrows():

        selector_items.append(
            {
                "id":
                    row["refund_id"],

                "label":
                    (
                        f"#{row['refund_id']} | "
                        f"{row['invoice_no']} | "
                        f"{money(row['report_total'])} | "
                        f"{row['status']}"
                    ),
            }
        )

    selected_refund_id = st.selectbox(
        "Select Refund",
        options=[
            item["id"]
            for item in selector_items
        ],
        format_func=lambda refund_id: next(
            (
                item["label"]
                for item in selector_items
                if item["id"]
                == refund_id
            ),
            str(refund_id),
        ),
    )

    # ==========================================================================
    # SELECTED ROW
    # ==========================================================================

    selected_rows = filtered[
        filtered["refund_id"]
        == selected_refund_id
    ]

    if selected_rows.empty:

        st.warning(
            "Selected refund was not found."
        )

    else:

        selected = (
            selected_rows.iloc[0]
        )

        # ======================================================================
        # HEADER
        # ======================================================================

        st.markdown(
            f"### Refund #{selected_refund_id}"
        )

        h1, h2, h3, h4 = st.columns(4)

        with h1:

            st.caption(
                "Invoice"
            )

            st.write(
                selected[
                    "invoice_no"
                ]
                or "-"
            )

        with h2:

            st.caption(
                "Status"
            )

            st.write(
                selected[
                    "status"
                ]
            )

        with h3:

            st.caption(
                "Cashier"
            )

            st.write(
                selected[
                    "cashier_name"
                ]
                or "-"
            )

        with h4:

            st.caption(
                "Warehouse"
            )

            st.write(
                selected[
                    "warehouse_name"
                ]
                or "-"
            )

        # ======================================================================
        # DATE / REASON
        # ======================================================================

        d1, d2 = st.columns(2)

        with d1:

            st.caption(
                "Refund Date"
            )

            if pd.notna(
                selected[
                    "refund_date"
                ]
            ):

                st.write(
                    selected[
                        "refund_date"
                    ].strftime(
                        "%Y-%m-%d %H:%M"
                    )
                )

            else:

                st.write("-")

        with d2:

            st.caption(
                "Reason"
            )

            st.write(
                selected[
                    "reason"
                ]
                or "-"
            )

        # ======================================================================
        # SELECTED STATUS FINANCIAL LOGIC
        # ======================================================================

        selected_status = (
            safe_text(
                selected["status"]
            )
            .upper()
        )

        if selected_status in (
            ACTUAL_REFUND_STATUSES
        ):

            selected_net = (
                safe_float(
                    selected[
                        "report_net"
                    ]
                )
            )

            selected_tax = (
                safe_float(
                    selected[
                        "report_tax"
                    ]
                )
            )

            selected_total = (
                safe_float(
                    selected[
                        "report_total"
                    ]
                )
            )

        elif selected_status == REJECTED_STATUS:

            selected_net = 0.0
            selected_tax = 0.0
            selected_total = 0.0

        else:

            selected_net = 0.0
            selected_tax = 0.0
            selected_total = 0.0

        # ======================================================================
        # AMOUNTS
        # ======================================================================

        st.divider()

        a1, a2, a3 = st.columns(3)

        with a1:

            st.metric(
                "Refund Net",
                money(selected_net),
            )

        with a2:

            st.metric(
                "Refund Tax",
                money(selected_tax),
            )

        with a3:

            st.metric(
                "Refund Total",
                money(selected_total),
            )

        # ======================================================================
        # STATUS NOTICE
        # ======================================================================

        if selected_status == REJECTED_STATUS:

            st.error(
                "❌ This refund is REJECTED. "
                "Its amount is NOT included in Refund Total."
            )

        elif selected_status == PENDING_STATUS:

            st.warning(
                "⏳ This refund is PENDING. "
                "Its amount is NOT included in Refund Total."
            )

        # ======================================================================
        # LOAD DETAIL
        # ======================================================================

        try:

            response = (
                db()
                .table(
                    "refund_detail_view"
                )
                .select("*")
                .eq(
                    "refund_id",
                    selected_refund_id,
                )
                .execute()
            )

            selected_items = (
                response.data or []
            )

        except Exception as e:

            selected_items = []

            st.error(
                f"Unable to load refund details: {e}"
            )

        # ======================================================================
        # DETAIL
        # ======================================================================

        if selected_items:

            detail_df = pd.DataFrame(
                selected_items
            )

            # ==================================================================
            # SAFE COLUMNS
            # ==================================================================

            for col in [
                "quantity",
                "unit_price",
                "item_total",
                "refund_net_amount",
                "refund_tax_amount",
                "refund_total_amount",
            ]:

                if col not in detail_df.columns:

                    detail_df[col] = 0

                detail_df[col] = pd.to_numeric(
                    detail_df[col],
                    errors="coerce",
                ).fillna(0)

            if (
                "product_name"
                not in detail_df.columns
            ):

                detail_df[
                    "product_name"
                ] = ""

            # ==================================================================
            # LEGACY NET
            # ==================================================================

            detail_df[
                "display_net"
            ] = detail_df[
                "refund_net_amount"
            ]

            legacy_detail_net = (
                (
                    detail_df[
                        "display_net"
                    ]
                    == 0
                )
                &
                (
                    detail_df[
                        "item_total"
                    ]
                    != 0
                )
            )

            detail_df.loc[
                legacy_detail_net,
                "display_net",
            ] = detail_df.loc[
                legacy_detail_net,
                "item_total",
            ]

            # ==================================================================
            # TAX
            # ==================================================================

            detail_df[
                "display_tax"
            ] = detail_df[
                "refund_tax_amount"
            ]

            # ==================================================================
            # TOTAL
            # ==================================================================

            detail_df[
                "display_total"
            ] = detail_df[
                "refund_total_amount"
            ]

            legacy_detail_total = (
                (
                    detail_df[
                        "display_total"
                    ]
                    == 0
                )
                &
                (
                    detail_df[
                        "item_total"
                    ]
                    != 0
                )
            )

            detail_df.loc[
                legacy_detail_total,
                "display_total",
            ] = detail_df.loc[
                legacy_detail_total,
                "item_total",
            ]

            # ==================================================================
            # IMPORTANT:
            # REJECTED / PENDING DISPLAY
            #
            # They can show original requested amount in detail,
            # but actual Refund Total KPI remains zero.
            # ==================================================================

            detail_display = pd.DataFrame(
                {
                    "Product":
                        detail_df[
                            "product_name"
                        ],

                    "Qty":
                        detail_df[
                            "quantity"
                        ],

                    "Unit Price":
                        detail_df[
                            "unit_price"
                        ],

                    "Refund Net":
                        detail_df[
                            "display_net"
                        ],

                    "Refund Tax":
                        detail_df[
                            "display_tax"
                        ],

                    "Refund Total":
                        detail_df[
                            "display_total"
                        ],
                }
            )

            st.subheader(
                "📦 Refund Items"
            )

            st.dataframe(
                detail_display,
                use_container_width=True,
                hide_index=True,

                column_config={

                    "Qty":
                        st.column_config.NumberColumn(
                            "Qty",
                            format="%.2f",
                        ),

                    "Unit Price":
                        st.column_config.NumberColumn(
                            "Unit Price",
                            format="%,.2f MMK",
                        ),

                    "Refund Net":
                        st.column_config.NumberColumn(
                            "Refund Net",
                            format="%,.2f MMK",
                        ),

                    "Refund Tax":
                        st.column_config.NumberColumn(
                            "Refund Tax",
                            format="%,.2f MMK",
                        ),

                    "Refund Total":
                        st.column_config.NumberColumn(
                            "Refund Total",
                            format="%,.2f MMK",
                        ),
                },
            )

            # ==================================================================
            # DETAIL TOTAL
            # ==================================================================

            detail_net = (
                detail_df[
                    "display_net"
                ].sum()
            )

            detail_tax = (
                detail_df[
                    "display_tax"
                ].sum()
            )

            detail_total = (
                detail_df[
                    "display_total"
                ].sum()
            )

            # ------------------------------------------------------------------
            # For REJECTED / PENDING:
            # financial actual total must remain zero.
            # ------------------------------------------------------------------

            if selected_status not in (
                ACTUAL_REFUND_STATUSES
            ):

                detail_actual_net = 0.0
                detail_actual_tax = 0.0
                detail_actual_total = 0.0

            else:

                detail_actual_net = (
                    detail_net
                )

                detail_actual_tax = (
                    detail_tax
                )

                detail_actual_total = (
                    detail_total
                )

            t1, t2, t3 = st.columns(3)

            with t1:

                st.metric(
                    "Actual Refund Net",
                    money(
                        detail_actual_net
                    ),
                )

            with t2:

                st.metric(
                    "Actual Refund Tax",
                    money(
                        detail_actual_tax
                    ),
                )

            with t3:

                st.metric(
                    "Actual Refund Total",
                    money(
                        detail_actual_total
                    ),
                )

        else:

            st.warning(
                "No refund item records found."
            )


# ==============================================================================
# ANALYTICS
# ==============================================================================

st.divider()

st.subheader(
    "📈 Refund Analytics"
)


if filtered.empty:

    st.info(
        "No data for analytics."
    )

else:

    # ==========================================================================
    # ACTUAL REFUND DAILY
    #
    # REJECTED / PENDING EXCLUDED
    # ==========================================================================

    actual_filtered = filtered[
        actual_refund_mask
    ].copy()

    if not actual_filtered.empty:

        daily_actual = (
            actual_filtered
            .assign(
                report_day=
                actual_filtered[
                    "refund_date"
                ].dt.date
            )
            .groupby(
                "report_day"
            )[
                "report_total"
            ]
            .sum()
            .sort_index()
        )

        st.markdown(
            "#### 💰 Daily Actual Refund Total"
        )

        st.line_chart(
            daily_actual,
            use_container_width=True,
        )

    else:

        st.info(
            "No completed/approved refunds "
            "for the selected period."
        )

    # ==========================================================================
    # PRODUCT + CASHIER
    # ==========================================================================

    ac1, ac2 = st.columns(2)

    # ==========================================================================
    # TOP PRODUCTS
    # ==========================================================================

    with ac1:

        st.markdown(
            "#### 🏆 Top Refunded Products"
        )

        if not actual_filtered.empty:

            top_products = (
                actual_filtered
                .groupby(
                    "product_name"
                )[
                    "quantity"
                ]
                .sum()
                .sort_values(
                    ascending=False
                )
                .head(10)
            )

            st.bar_chart(
                top_products,
                use_container_width=True,
            )

        else:

            st.info(
                "No actual refund data."
            )

    # ==========================================================================
    # CASHIER
    # ==========================================================================

    with ac2:

        st.markdown(
            "#### 👤 Cashier Refund Total"
        )

        if not actual_filtered.empty:

            cashier_data = (
                actual_filtered
                .groupby(
                    "cashier_name"
                )[
                    "report_total"
                ]
                .sum()
                .sort_values(
                    ascending=False
                )
                .head(10)
            )

            st.bar_chart(
                cashier_data,
                use_container_width=True,
            )

        else:

            st.info(
                "No actual refund data."
            )


# ==============================================================================
# FINAL SUMMARY
# ==============================================================================

st.divider()

st.subheader(
    "🧾 Final Report Summary"
)


s1, s2, s3, s4 = st.columns(4)


with s1:

    st.metric(
        "Actual Refund Total",
        money(total_refund),
    )


with s2:

    st.metric(
        "Rejected Total",
        money(rejected_total),
    )


with s3:

    st.metric(
        "Pending Total",
        money(pending_total),
    )


with s4:

    st.metric(
        "Report Records",
        f"{total_refunds:,}",
    )


# ==============================================================================
# ACCOUNTING NOTE
# ==============================================================================

st.caption(
    "Accounting Rule: "
    "Refund Total includes only COMPLETED and APPROVED refunds. "
    "REJECTED and PENDING amounts are excluded from Refund Total "
    "and reported separately."
)


# ==============================================================================
# END
# ==============================================================================
